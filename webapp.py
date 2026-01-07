"""
Kasboek Debutade - Web Applicatie
===================================

Een moderne web-gebaseerde applicatie voor het beheren van kasboektransacties.
Dit is de Flask web app versie van de originele Tkinter applicatie.

Functionaliteiten:
- Invoeren van financiële gegevens via een webinterface
- Validatie van invoer (datums en bedragen)
- Automatische opslag in Excel-bestand
- Logging van gebeurtenissen
- Overzicht van recente transacties
- Berekening van totaal kassaldo

Versie: 2.0 (Web App)
Datum: 2026-01-03
Auteur: Eric G.
"""

from flask import Flask, render_template, request, jsonify, redirect, url_for, send_from_directory
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import json
import logging
import shutil
import locale
import getpass
import sys

# Fix encoding voor Windows console
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except AttributeError:
        import io
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Vereiste kolom headers voor het Excel bestand
REQUIRED_HEADERS = [
    "Datum",
    "Naam / Omschrijving",
    "Rekening",
    "Tegenrekening",
    "Code",
    "Af Bij",
    "Bedrag (EUR)",
    "Mutatiesoort",
    "Mededelingen",
    "Saldo na mutatie",
    "",
    "Tag"
]

app = Flask(__name__)
app.static_folder = 'static'

# Laad configuratie
def load_config(config_path):
    """Laad de configuratie uit een JSON bestand"""
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"Configuratiebestand niet gevonden: {config_path}")
    
    with open(config_path, "r", encoding="utf-8") as config_file:
        config = json.load(config_file)
    
    required_keys = ["excel_file_path", "resources", 
                    "backup_directory", "log_directory", "excel_sheet_name", 
                    "tags", "log_level"]
    
    for key in required_keys:
        if key not in config:
            raise KeyError(f"Configuratiesleutel ontbreekt: {key}")
    
    return config

def save_config(config_data, config_path=None):
    """Sla configuratie op naar JSON bestand"""
    target_path = config_path or CONFIG_PATH
    try:
        with open(target_path, "w", encoding="utf-8") as config_file:
            json.dump(config_data, config_file, indent=4)
        return True
    except Exception as e:
        logging.error(f"Fout bij opslaan configuratie: {str(e)}")
        return False


def validate_excel_headers(file_path, required_headers=REQUIRED_HEADERS):
    """Controleer of de Excel headers overeenkomen met het vereiste formaat"""
    wb = None
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheet = wb[EXCEL_SHEET_NAME] if EXCEL_SHEET_NAME in wb.sheetnames else wb.active
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        normalized_row = [str(val).strip() if val is not None else '' for val in first_row]
        # Pad rijen zodat lengte gelijk is voor vergelijking
        if len(normalized_row) < len(required_headers):
            normalized_row += [''] * (len(required_headers) - len(normalized_row))
        normalized_required = [str(val).strip() for val in required_headers]
        return normalized_row[:len(normalized_required)] == normalized_required
    except Exception as e:
        logging.error(f"Fout bij valideren Excel headers: {str(e)}")
        return False
    finally:
        if wb:
            wb.close()

# Bepaal het directory waar het script zich bevindt
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Standaard configuratie pad (kan worden aangepast via omgevingsvariabele)
CONFIG_PATH = os.getenv('KASBOEK_CONFIG', 
    os.path.join(SCRIPT_DIR, 'config.json'))

try:
    config = load_config(CONFIG_PATH)
except (FileNotFoundError, KeyError) as e:
    print(f"WAARSCHUWING: {e}")
    # Gebruik standaard configuratie voor ontwikkeling
    config = {
        "excel_file_directory": r"C:\Users\ericg\OneDrive\Documents\Code",
        "excel_file_name": "records.xlsx",
        "resources": r"C:\Users\ericg\OneDrive\Documents\Code\resources",
        "backup_directory": r"C:\Users\ericg\OneDrive\Documents\Code\backups",
        "log_directory": r"C:\Users\ericg\OneDrive\Documents\Code\logs",
        "excel_sheet_name": "Transacties",
        "tags": ["Algemeen", "Evenement", "Materiaal", "Training", "Overig"],
        "log_level": "INFO"
    }

# Configuratie variabelen
EXCEL_FILE_PATH = config["excel_file_path"]
EXCEL_FILE_DIRECTORY = os.path.dirname(EXCEL_FILE_PATH)
EXCEL_FILE_NAME = os.path.basename(EXCEL_FILE_PATH)
BACKUP_DIRECTORY = config["backup_directory"]
LOG_DIRECTORY = config["log_directory"]
EXCEL_SHEET_NAME = config["excel_sheet_name"]
TAGS = config["tags"]
LOG_LEVEL = config["log_level"]

# Valideer alle bestandspaden bij startup
def validate_config():
    """Valideer configuratie - start wel maar waarschuw als Excel pad leeg is"""
    
    logging.info("=" * 70)
    logging.info("CONFIGURATIE VALIDATIE")
    logging.info(f"Excel pad uit config: '{EXCEL_FILE_PATH}'")
    
    # Controleer of Excel pad leeg is
    if not EXCEL_FILE_PATH or EXCEL_FILE_PATH.strip() == "":
        logging.warning("Excel bestandspad is LEEG")
        logging.warning("Gebruiker MOET eerst via Instellingen een bestand selecteren!")
        return True  # App start wel, maar gebruiker moet eerst Excel bestand kiezen
    else:
        logging.info(f"Bestand bestaat: {os.path.exists(EXCEL_FILE_PATH)}")
        
        # Controleer of Excel bestand bestaat
        if not os.path.exists(EXCEL_FILE_PATH):
            logging.warning(f"Excel bestand niet gevonden: {EXCEL_FILE_PATH}")
            # Extra debugging
            dir_path = os.path.dirname(EXCEL_FILE_PATH)
            if os.path.exists(dir_path):
                logging.info(f"Directory bestaat wel: {dir_path}")
                try:
                    files = os.listdir(dir_path)
                    xlsx_files = [f for f in files if f.endswith('.xlsx')]
                    logging.info(f".xlsx bestanden in directory: {xlsx_files}")
                except:
                    pass
            else:
                logging.warning(f"Directory bestaat niet: {dir_path}")
            return True  # App start wel
        else:
            logging.info(f"Excel bestand gevonden: {EXCEL_FILE_PATH}")
    
    # Controleer of directories bestaan, anders aanmaken
    for dir_name, dir_path in [("Backup", BACKUP_DIRECTORY), ("Log", LOG_DIRECTORY), ("Excel", EXCEL_FILE_DIRECTORY)]:
        if not os.path.exists(dir_path):
            try:
                os.makedirs(dir_path)
                logging.info(f"{dir_name} directory aangemaakt: {dir_path}")
            except Exception as e:
                logging.error(f"Kan {dir_name} directory niet aanmaken: {dir_path}")
        else:
            logging.info(f"{dir_name} directory bestaat: {dir_path}")
    
    # Controleer of log file schrijfbaar is
    if os.path.exists(LOG_DIRECTORY):
        try:
            test_log = os.path.join(LOG_DIRECTORY, ".write_test")
            with open(test_log, "w") as f:
                f.write("test")
            os.remove(test_log)
            logging.info("Log directory is schrijfbaar")
        except Exception as e:
            logging.error(f"Log directory is niet schrijfbaar: {str(e)}")
    
    logging.info("Applicatie start!")
    logging.info("=" * 70)
    return True

# Stel Nederlandse locale in (optioneel, kan problemen geven op sommige systemen)
try:
    locale.setlocale(locale.LC_TIME, "nl_NL")
except:
    pass

# Maak backup bij opstarten
def create_backup():
    """Maak een backup van het Excel bestand"""
    try:
        if os.path.exists(EXCEL_FILE_PATH):
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = os.path.join(BACKUP_DIRECTORY, 
                f"{EXCEL_FILE_NAME}_backup_{timestamp}.xlsx")
            shutil.copy(EXCEL_FILE_PATH, backup_path)
            logging.info(f"Backup gemaakt: {backup_path}")
            return True
    except Exception as e:
        logging.error(f"Fout bij maken backup: {str(e)}")
        return False

def calculate_total_amount():
    """Bereken het totale saldo in de kas"""
    try:
        if not os.path.exists(EXCEL_FILE_PATH):
            return 0
        
        wb = load_workbook(EXCEL_FILE_PATH)
        if EXCEL_SHEET_NAME in wb.sheetnames:
            sheet = wb[EXCEL_SHEET_NAME]
            total = 0
            # Kolom F = Af/Bij (kolom 6), Kolom G = Bedrag (kolom 7)
            for row in sheet.iter_rows(min_row=2, min_col=6, max_col=7, values_only=True):
                af_bij, amount = row
                if isinstance(amount, (int, float)):
                    if af_bij == "Af":
                        total -= amount
                    elif af_bij == "Bij":
                        total += amount
            return round(total, 2)
        return 0
    except Exception as e:
        logging.error(f"Fout bij berekenen totaal: {str(e)}")
        return 0

def get_recent_transactions(limit=10):
    """Haal de meest recente transacties op"""
    try:
        if not os.path.exists(EXCEL_FILE_PATH):
            return []
        
        wb = load_workbook(EXCEL_FILE_PATH)
        if EXCEL_SHEET_NAME not in wb.sheetnames:
            return []
        
        sheet = wb[EXCEL_SHEET_NAME]
        transactions = []
        
        # Start bij rij 2 (rij 1 is header)
        for row in sheet.iter_rows(min_row=2, max_row=min(limit+1, sheet.max_row), 
                                   values_only=True):
            if row[0]:  # Als datum bestaat
                transactions.append({
                    'datum': row[0].strftime('%Y-%m-%d') if isinstance(row[0], datetime) else str(row[0]),
                    'mededelingen': (row[8] if len(row) > 8 else None) or row[1] or '',
                    'af_bij': row[5] or '',
                    'bedrag': f"{row[6]:.2f}" if isinstance(row[6], (int, float)) else '0.00',
                    'tag': row[11] or '',
                    'saldo': f"€ {row[9]:.2f}" if isinstance(row[9], (int, float)) else '€ 0.00'
                })
        
        return transactions
    except Exception as e:
        logging.error(f"Fout bij ophalen transacties: {str(e)}")
        return []

def get_all_transactions():
    """Haal alle transacties op uit het Excel bestand"""
    try:
        if not os.path.exists(EXCEL_FILE_PATH):
            return []
        
        wb = load_workbook(EXCEL_FILE_PATH)
        if EXCEL_SHEET_NAME not in wb.sheetnames:
            return []
        
        sheet = wb[EXCEL_SHEET_NAME]
        transactions = []
        
        # Start bij rij 2 (rij 1 is header)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Als datum bestaat
                transactions.append({
                    'datum': row[0].strftime('%Y-%m-%d') if isinstance(row[0], datetime) else str(row[0]),
                    'mededelingen': (row[8] if len(row) > 8 else None) or row[1] or '',
                    'af_bij': row[5] or '',
                    'bedrag': f"{row[6]:.2f}" if isinstance(row[6], (int, float)) else '0.00',
                    'rekening': row[2] or '',
                    'tag': row[11] or ''
                })
        
        return transactions
    except Exception as e:
        logging.error(f"Fout bij ophalen alle transacties: {str(e)}")
        return []

@app.route('/favicon.ico')
def favicon():
    """Serve the favicon"""
    return send_from_directory(app.static_folder, 'icon.ico', mimetype='image/vnd.microsoft.icon')

@app.route('/')
def index():
    """Hoofdpagina met invoerformulier"""
    total_amount = calculate_total_amount()
    recent_transactions = get_recent_transactions()
    today = datetime.now().strftime('%Y-%m-%d')
    current_date_display = datetime.now().strftime('%d-%m-%Y')
    current_user = getpass.getuser()
    
    return render_template('index.html', 
                         tags=TAGS,
                         total_amount=total_amount,
                         recent_transactions=recent_transactions,
                         today=today,
                         current_date=current_date_display,
                         current_user=current_user)

@app.route('/add_transaction', methods=['POST'])
def add_transaction():
    """Voeg een nieuwe transactie toe"""
    try:
        # Controleer of Excel bestand is ingesteld en bestaat
        if not EXCEL_FILE_PATH or EXCEL_FILE_PATH.strip() == "":
            return jsonify({
                'success': False, 
                'message': 'Geen Excel bestand ingesteld. Ga naar Instellingen en selecteer/upload een Excel bestand.'
            }), 400
        
        if not os.path.exists(EXCEL_FILE_PATH):
            return jsonify({
                'success': False, 
                'message': 'Excel bestand niet gevonden. Ga naar Instellingen en selecteer/upload een geldig Excel bestand.'
            }), 400

        # Haal gegevens op uit het formulier
        data = {
            'datum': request.form.get('datum'),
            'mededelingen': request.form.get('mededelingen', ''),
            'rekening': request.form.get('rekening', ''),
            'tegenrekening': request.form.get('tegenrekening', ''),
            'code': request.form.get('code', ''),
            'af_bij': request.form.get('af_bij'),
            'bedrag': request.form.get('bedrag'),
            'mutatiesoort': request.form.get('mutatiesoort', 'Kas'),
            'saldo': request.form.get('saldo', ''),
            'tag': request.form.get('tag', '')
        }
        
        # Validatie
        if not data['datum']:
            return jsonify({'success': False, 'message': 'Datum is verplicht'}), 400
        
        if not data['mededelingen'].strip():
            return jsonify({'success': False, 'message': 'Mededeling is verplicht'}), 400
        
        if not data['bedrag'].strip():
            return jsonify({'success': False, 'message': 'Bedrag is verplicht'}), 400
        
        # Converteer bedrag (accepteer komma als decimaal scheidingsteken)
        try:
            bedrag = float(data['bedrag'].replace(',', '.'))
        except ValueError:
            return jsonify({'success': False, 'message': 'Ongeldig bedrag'}), 400
        
        # Parse datum
        try:
            datum = datetime.strptime(data['datum'], '%Y-%m-%d')
        except ValueError:
            return jsonify({'success': False, 'message': 'Ongeldige datum'}), 400
        
        # Laad of maak Excel bestand
        if os.path.exists(EXCEL_FILE_PATH):
            wb = load_workbook(EXCEL_FILE_PATH)
            if EXCEL_SHEET_NAME in wb.sheetnames:
                sheet = wb[EXCEL_SHEET_NAME]
            else:
                sheet = wb.create_sheet(EXCEL_SHEET_NAME)
        else:
            wb = Workbook()
            sheet = wb.active
            sheet.title = EXCEL_SHEET_NAME
            # Voeg headers toe
            headers = ['Datum', 'Naam/Omschrijving', 'Rekening', 'Tegen Rekening', 
                      'Code', 'Af Bij', 'Bedrag', 'Mutatiesoort', 'Mededelingen', 
                      'Saldo na mutatie', '', 'Tag']
            sheet.append(headers)
        
        # Voeg lege rij in op positie 2
        sheet.insert_rows(2)
        
        # Voeg data toe op rij 2
        row_data = [
            datum,
            data['mededelingen'],
            data['rekening'],
            data['tegenrekening'],
            data['code'],
            data['af_bij'],
            bedrag,
            data['mutatiesoort'],
            data['mededelingen'],
            data['saldo'],
            '',
            data['tag']
        ]
        
        for col, value in enumerate(row_data, start=1):
            sheet.cell(row=2, column=col, value=value)
        
        # Sla op
        wb.save(EXCEL_FILE_PATH)
        
        # Log de actie met meer details
        user = getpass.getuser()  # Krijg Windows username
        ip_addr = request.remote_addr  # IP adres
        logging.info(f"TRANSACTIE TOEGEVOEGD | Gebruiker: {user} | IP: {ip_addr} | Datum: {data['datum']} | "
                    f"Beschrijving: {data['mededelingen']} | Bedrag: €{bedrag} | Af/Bij: {data['af_bij']} | Tag: {data['tag']}")
        
        # Bereken nieuw totaal
        new_total = calculate_total_amount()
        
        return jsonify({
            'success': True, 
            'message': 'Transactie succesvol opgeslagen!',
            'new_total': new_total
        })
        
    except Exception as e:
        logging.error(f"Fout bij toevoegen transactie: {str(e)}")
        return jsonify({'success': False, 'message': f'Fout: {str(e)}'}), 500

@app.route('/api/recommend-category', methods=['POST'])
def recommend_category():
    """Geef categorie aanbevelingen op basis van de mededelingen tekst"""
    try:
        data = request.get_json() or {}
        description = str(data.get('description', '')).strip().lower()
        
        if not description or len(description) < 3:
            return jsonify({'recommendations': []})
        
        # Laad de test set
        test_set_path = os.path.join('static', 'category_test_set.xlsx')
        if not os.path.exists(test_set_path):
            logging.warning("Category test set niet gevonden")
            return jsonify({'recommendations': []})
        
        wb = load_workbook(test_set_path)
        sheet = wb.active
        
        # Verzamel alle mededelingen en categorieën
        training_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[2] and row[3]:  # Mededelingen en Tag kolommen
                training_data.append({
                    'description': str(row[2]).lower(),
                    'category': str(row[3])
                })
        
        wb.close()
        
        # Bereken similarity scores
        def calculate_similarity(text1, text2):
            """Simpele similarity based op woord overlap"""
            words1 = set(text1.split())
            words2 = set(text2.split())
            if not words1 or not words2:
                return 0
            intersection = words1.intersection(words2)
            union = words1.union(words2)
            return len(intersection) / len(union)
        
        # Vind matches
        matches = []
        for item in training_data:
            score = calculate_similarity(description, item['description'])
            if score > 0:
                matches.append({
                    'category': item['category'],
                    'score': score,
                    'example': item['description']
                })
        
        # Sorteer op score en groepeer per categorie
        matches.sort(key=lambda x: x['score'], reverse=True)
        
        # Houd unieke categorieën bij met hoogste score
        seen_categories = {}
        for match in matches:
            cat = match['category']
            if cat not in seen_categories or match['score'] > seen_categories[cat]['score']:
                seen_categories[cat] = match
        
        # Neem top 5 aanbevelingen
        recommendations = sorted(seen_categories.values(), 
                               key=lambda x: x['score'], 
                               reverse=True)[:5]
        
        return jsonify({'recommendations': recommendations})
        
    except Exception as e:
        logging.error(f"Fout bij category recommendation: {str(e)}")
        return jsonify({'recommendations': []})

@app.route('/get_total')
def get_total():
    """Haal het huidige totaal op"""
    total = calculate_total_amount()
    return jsonify({'total': total})

@app.route('/get_transactions')
def get_transactions():
    """Haal recente transacties op (AJAX)"""
    transactions = get_recent_transactions()
    return jsonify({'transactions': transactions})

@app.route('/api/all_transactions')
def api_all_transactions():
    """Haal alle transacties op (AJAX) voor de history"""
    transactions = get_all_transactions()
    return jsonify({'transactions': transactions})

@app.route('/backup')
def backup():
    """Maak handmatig een backup"""
    success = create_backup()
    if success:
        return jsonify({'success': True, 'message': 'Backup succesvol gemaakt'})
    else:
        return jsonify({'success': False, 'message': 'Fout bij maken backup'}), 500

@app.route('/quit', methods=['POST'])
def quit_application():
    """Stop de applicatie en log dit"""
    try:
        user = getpass.getuser()
        ip_addr = request.remote_addr
        duration = request.get_json().get('duration', 'Onbekend') if request.is_json else 'Onbekend'
        
        logging.info(f"APPLICATIE AFGESLOTEN | Gebruiker: {user} | IP: {ip_addr} | Sessieduur: {duration}")
        logging.info("=" * 70)
        
        # Stuur succes response terug naar client
        response = jsonify({'success': True, 'message': 'Applicatie sluit af'})
        
        # Schedule de shutdown na een korte vertraging zodat response kan worden verzonden
        def shutdown_server():
            import time
            time.sleep(1)  # Wacht 1 seconde zodat response verzonden kan worden
            logging.info("Flask server wordt beëindigd...")
            os._exit(0)
        
        import threading
        shutdown_thread = threading.Thread(target=shutdown_server, daemon=True)
        shutdown_thread.start()
        
        return response, 200
    except Exception as e:
        logging.error(f"Fout bij afsluiten applicatie: {str(e)}")
        return jsonify({'success': False, 'message': f'Fout: {str(e)}'}), 500

@app.route('/settings')
def settings():
    """Toon instellingen pagina"""
    user = getpass.getuser()
    ip_addr = request.remote_addr
    logging.info(f"INSTELLINGEN GEOPEND | Gebruiker: {user} | IP: {ip_addr}")
    
    current_date_display = datetime.now().strftime('%d-%m-%Y')
    current_user = getpass.getuser()
    
    settings_info = {
        'excel_file_name': EXCEL_FILE_NAME,
        'excel_file_directory': EXCEL_FILE_DIRECTORY,
        'excel_file_path': EXCEL_FILE_PATH,
        'backup_directory': BACKUP_DIRECTORY,
        'log_directory': LOG_DIRECTORY,
        'backup_dir': BACKUP_DIRECTORY,
        'log_dir': LOG_DIRECTORY,
        'sheet_name': EXCEL_SHEET_NAME,
        'log_level': LOG_LEVEL,
        'tags': TAGS
    }
    return render_template('settings.html', settings=settings_info, current_date=current_date_display, current_user=current_user)

@app.route('/settings/excel-file', methods=['POST'])
def update_excel_file():
    """Werk het Excel bestandspad bij en sla configuratie op"""
    try:
        user = getpass.getuser()
        ip_addr = request.remote_addr
        data = request.get_json() or {}
        new_name = str(data.get('excel_file_name', '')).strip()

        if not new_name:
            return jsonify({'success': False, 'message': 'Bestandsnaam is verplicht'}), 400

        if not new_name.lower().endswith('.xlsx'):
            new_name = f"{new_name}.xlsx"

        global EXCEL_FILE_NAME, EXCEL_FILE_PATH, config
        old_path = EXCEL_FILE_PATH
        EXCEL_FILE_NAME = new_name
        EXCEL_FILE_PATH = os.path.join(EXCEL_FILE_DIRECTORY, EXCEL_FILE_NAME)
        config['excel_file_path'] = EXCEL_FILE_PATH

        if not save_config(config):
            return jsonify({'success': False, 'message': 'Opslaan in config.json is mislukt'}), 500

        logging.info(f"INSTELLING GEWIJZIGD | Gebruiker: {user} | IP: {ip_addr} | Setting: Excel bestandsnaam | "
                    f"Van: {old_path} | Naar: {EXCEL_FILE_PATH}")

        return jsonify({
            'success': True,
            'excel_file_name': EXCEL_FILE_NAME,
            'excel_file_path': EXCEL_FILE_PATH
        })
    except Exception as e:
        logging.error(f"Fout bij bijwerken excel bestandspad: {str(e)}")
        return jsonify({'success': False, 'message': f'Fout: {str(e)}'}), 500


@app.route('/settings/excel-file-path', methods=['POST'])
def set_excel_file_path():
    """Stel direct een bestaand Excel pad in zonder kopieeren"""
    try:
        user = getpass.getuser()
        ip_addr = request.remote_addr
        data = request.get_json() or {}
        new_path = str(data.get('excel_file_path', '')).strip()

        if not new_path:
            return jsonify({'success': False, 'message': 'Pad is verplicht'}), 400

        if not new_path.lower().endswith('.xlsx'):
            return jsonify({'success': False, 'message': 'Bestand moet een .xlsx zijn'}), 400

        if not os.path.exists(new_path):
            return jsonify({'success': False, 'message': 'Bestand niet gevonden op opgegeven pad'}), 400

        if not validate_excel_headers(new_path):
            return jsonify({'success': False, 'message': 'Bestand voldoet niet aan het vereiste kolom formaat'}), 400

        global EXCEL_FILE_NAME, EXCEL_FILE_PATH, EXCEL_FILE_DIRECTORY, config
        old_path = EXCEL_FILE_PATH
        EXCEL_FILE_PATH = new_path
        EXCEL_FILE_NAME = os.path.basename(new_path)
        EXCEL_FILE_DIRECTORY = os.path.dirname(new_path)
        config['excel_file_path'] = EXCEL_FILE_PATH

        if not save_config(config):
            return jsonify({'success': False, 'message': 'Opslaan in config.json is mislukt'}), 500

        logging.info(f"INSTELLING GEWIJZIGD | Gebruiker: {user} | IP: {ip_addr} | Setting: Excel bestandspad | "
                    f"Van: {old_path} | Naar: {EXCEL_FILE_PATH}")

        return jsonify({
            'success': True,
            'excel_file_name': EXCEL_FILE_NAME,
            'excel_file_path': EXCEL_FILE_PATH
        })
    except Exception as e:
        logging.error(f"Fout bij instellen excel pad: {str(e)}")
        return jsonify({'success': False, 'message': f'Fout: {str(e)}'}), 500

@app.route('/settings/excel-file-upload', methods=['POST'])
def upload_excel_file():
    """Upload een Excel bestand, sla het op in dezelfde directory en werk config bij"""
    try:
        user = getpass.getuser()
        ip_addr = request.remote_addr
        
        if 'excel_file' not in request.files:
            return jsonify({'success': False, 'message': 'Geen bestand ontvangen'}), 400

        file = request.files['excel_file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'Geen bestand geselecteerd'}), 400

        filename = secure_filename(file.filename)
        if not filename.lower().endswith('.xlsx'):
            return jsonify({'success': False, 'message': 'Alleen .xlsx bestanden zijn toegestaan'}), 400

        save_path = os.path.join(EXCEL_FILE_DIRECTORY, filename)
        file.save(save_path)

        # Valideer kolom headers
        if not validate_excel_headers(save_path):
            os.remove(save_path)
            return jsonify({
                'success': False,
                'message': 'Het gekozen Excel bestand voldoet niet aan het vereiste formaat (ontbrekende of onjuiste kolom headers).'
            }), 400

        global EXCEL_FILE_NAME, EXCEL_FILE_PATH, config
        old_path = EXCEL_FILE_PATH
        EXCEL_FILE_NAME = filename
        EXCEL_FILE_PATH = save_path
        config['excel_file_path'] = EXCEL_FILE_PATH

        if not save_config(config):
            return jsonify({'success': False, 'message': 'Opslaan in config.json is mislukt'}), 500

        logging.info(f"INSTELLING GEWIJZIGD | Gebruiker: {user} | IP: {ip_addr} | Setting: Excel bestand geupload | "
                    f"Van: {old_path} | Naar: {EXCEL_FILE_PATH} | Bestand: {filename}")

        return jsonify({
            'success': True,
            'excel_file_name': EXCEL_FILE_NAME,
            'excel_file_path': EXCEL_FILE_PATH
        })
    except Exception as e:
        logging.error(f"Fout bij uploaden excel bestand: {str(e)}")
        return jsonify({'success': False, 'message': f'Fout: {str(e)}'}), 500

@app.route('/settings/backup-directory', methods=['POST'])
def set_backup_directory():
    """Stel backup directory pad in en sla op in config"""
    try:
        user = getpass.getuser()
        ip_addr = request.remote_addr
        data = request.get_json() or {}
        new_path = str(data.get('backup_directory', '')).strip()

        if not new_path:
            return jsonify({'success': False, 'message': 'Pad is verplicht'}), 400

        global BACKUP_DIRECTORY, config
        old_path = BACKUP_DIRECTORY
        BACKUP_DIRECTORY = new_path
        config['backup_directory'] = BACKUP_DIRECTORY

        if not save_config(config):
            return jsonify({'success': False, 'message': 'Opslaan in config.json is mislukt'}), 500

        logging.info(f"INSTELLING GEWIJZIGD | Gebruiker: {user} | IP: {ip_addr} | Setting: Backup directory | "
                    f"Van: {old_path} | Naar: {BACKUP_DIRECTORY}")

        return jsonify({
            'success': True,
            'backup_directory': BACKUP_DIRECTORY
        })
    except Exception as e:
        logging.error(f"Fout bij instellen backup directory: {str(e)}")
        return jsonify({'success': False, 'message': f'Fout: {str(e)}'}), 500


@app.route('/settings/log-directory', methods=['POST'])
def set_log_directory():
    """Stel log directory pad in en sla op in config"""
    try:
        user = getpass.getuser()
        ip_addr = request.remote_addr
        data = request.get_json() or {}
        new_path = str(data.get('log_directory', '')).strip()

        if not new_path:
            return jsonify({'success': False, 'message': 'Pad is verplicht'}), 400

        global LOG_DIRECTORY, config
        old_path = LOG_DIRECTORY
        LOG_DIRECTORY = new_path
        config['log_directory'] = LOG_DIRECTORY

        if not save_config(config):
            return jsonify({'success': False, 'message': 'Opslaan in config.json is mislukt'}), 500

        logging.info(f"INSTELLING GEWIJZIGD | Gebruiker: {user} | IP: {ip_addr} | Setting: Log directory | "
                    f"Van: {old_path} | Naar: {LOG_DIRECTORY}")

        return jsonify({
            'success': True,
            'log_directory': LOG_DIRECTORY
        })
    except Exception as e:
        logging.error(f"Fout bij instellen log directory: {str(e)}")
        return jsonify({'success': False, 'message': f'Fout: {str(e)}'}), 500


@app.route('/settings/log-level', methods=['POST'])
def set_log_level():
    """Stel log level in en sla op in config"""
    try:
        user = getpass.getuser()
        ip_addr = request.remote_addr
        data = request.get_json() or {}
        new_level = str(data.get('log_level', '')).strip().upper()

        valid_levels = ['DEBUG', 'INFO', 'WARNING', 'ERROR', 'CRITICAL']
        if new_level not in valid_levels:
            return jsonify({'success': False, 'message': 'Ongeldig log level'}), 400

        global LOG_LEVEL, config
        old_level = LOG_LEVEL
        LOG_LEVEL = new_level
        config['log_level'] = LOG_LEVEL

        if not save_config(config):
            return jsonify({'success': False, 'message': 'Opslaan in config.json is mislukt'}), 500

        # Pas het runtime log level meteen toe
        logging.getLogger().setLevel(getattr(logging, LOG_LEVEL))

        logging.info(f"INSTELLING GEWIJZIGD | Gebruiker: {user} | IP: {ip_addr} | Setting: Log level | "
                    f"Van: {old_level} | Naar: {LOG_LEVEL}")

        return jsonify({
            'success': True,
            'log_level': LOG_LEVEL
        })
    except Exception as e:
        logging.error(f"Fout bij instellen log level: {str(e)}")
        return jsonify({'success': False, 'message': f'Fout: {str(e)}'}), 500


@app.route('/settings/excel-sheet-name', methods=['POST'])
def set_excel_sheet_name():
    """Stel Excel sheet naam in met validatie"""
    try:
        user = getpass.getuser()
        ip_addr = request.remote_addr
        data = request.get_json() or {}
        new_sheet_name = str(data.get('sheet_name', '')).strip()

        if not new_sheet_name:
            return jsonify({'success': False, 'message': 'Sheet naam is verplicht'}), 400

        # Controleer of Excel bestand bestaat en de sheet naam daarin
        if not os.path.exists(EXCEL_FILE_PATH):
            return jsonify({'success': False, 'message': 'Excel bestand niet gevonden'}), 400

        # Controleer of de sheet bestaat in het Excel bestand
        try:
            wb = load_workbook(EXCEL_FILE_PATH, read_only=True, data_only=True)
            if new_sheet_name not in wb.sheetnames:
                available_sheets = ', '.join(wb.sheetnames) if wb.sheetnames else 'Geen sheets beschikbaar'
                wb.close()
                return jsonify({
                    'success': False,
                    'message': f'Sheet "{new_sheet_name}" niet gevonden in Excel bestand. Beschikbare sheets: {available_sheets}'
                }), 400
            
            # Controleer of de kolom headers juist zijn voor deze sheet
            sheet = wb[new_sheet_name]
            first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
            normalized_row = [str(val).strip() if val is not None else '' for val in first_row]
            
            # Pad rijen zodat lengte gelijk is voor vergelijking
            if len(normalized_row) < len(REQUIRED_HEADERS):
                normalized_row += [''] * (len(REQUIRED_HEADERS) - len(normalized_row))
            normalized_required = [str(val).strip() for val in REQUIRED_HEADERS]
            
            if normalized_row[:len(normalized_required)] != normalized_required:
                wb.close()
                return jsonify({
                    'success': False,
                    'message': f'Sheet "{new_sheet_name}" heeft onjuiste kolom headers. Verwacht: {", ".join(REQUIRED_HEADERS)}'
                }), 400
            
            wb.close()
        except jsonify as e:
            return e
        except Exception as e:
            logging.error(f"Fout bij controleren Excel sheet: {str(e)}")
            return jsonify({'success': False, 'message': f'Fout bij controleren Excel bestand: {str(e)}'}), 400

        global EXCEL_SHEET_NAME, config
        old_sheet_name = EXCEL_SHEET_NAME
        EXCEL_SHEET_NAME = new_sheet_name
        config['excel_sheet_name'] = EXCEL_SHEET_NAME

        if not save_config(config):
            return jsonify({'success': False, 'message': 'Opslaan in config.json is mislukt'}), 500

        logging.info(f"INSTELLING GEWIJZIGD | Gebruiker: {user} | IP: {ip_addr} | Setting: Excel sheet naam | "
                    f"Van: {old_sheet_name} | Naar: {EXCEL_SHEET_NAME}")

        return jsonify({
            'success': True,
            'sheet_name': EXCEL_SHEET_NAME
        })
    except Exception as e:
        logging.error(f"Fout bij instellen excel sheet naam: {str(e)}")
        return jsonify({'success': False, 'message': f'Fout: {str(e)}'}), 500


if __name__ == '__main__':
    print("=" * 60)
    print(">> Kasboek Debutade Web Applicatie - Startup")
    print("=" * 60)
    
    # Zorg dat log directory bestaat voordat we logging configureren
    if not os.path.exists(LOG_DIRECTORY):
        try:
            os.makedirs(LOG_DIRECTORY)
            print(f"Log directory aangemaakt: {LOG_DIRECTORY}")
        except Exception as e:
            print(f"FOUT: Kan log directory niet aanmaken: {LOG_DIRECTORY}")
            print(f"Details: {str(e)}")
            exit(1)
    
    # Configureer logging EERST zodat alle logs worden geschreven
    log_file_path = os.path.join(LOG_DIRECTORY, "kasboek_webapp_log.txt")
    logging.basicConfig(
        filename=log_file_path,
        level=getattr(logging, LOG_LEVEL.upper(), logging.INFO),
        format="%(asctime)s - %(levelname)s - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )
    
    # Valideer configuratie (maakt directories aan)
    if not validate_config():
        print("\n>> FOUT: Applicatie kan niet starten. Zorg dat config.json correct is ingesteld.")
        exit(1)
    
    # Maak backup bij starten
    create_backup()
    
    # Log startup met gebruikersinfo
    user = getpass.getuser()
    logging.info("=" * 70)
    logging.info(f"APPLICATIE GESTART | Gebruiker: {user} | Tijd: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logging.info("=" * 70)
    
    print("\n>> Applicatie is klaar om te starten!")
    print("=" * 60)
    
    # Start de Flask applicatie
    # Debug=False voor productie, True voor ontwikkeling
    # host='0.0.0.0' maakt de server toegankelijk vanaf andere apparaten op het netwerk
    app.run(debug=False, host='0.0.0.0', port=5000)
